"""
excel_manager.py — Lógica del Excel maestro (registro_maestro.xlsx).
Gestiona la creación, carga, backup y escritura del archivo maestro (Esto es en base al formato de excel requerido por el usuario, en este caso en como trabaja mi mamá).
"""

import os
import shutil
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# Ruta del Excel maestro
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
MASTER_FILE = os.path.join(DATA_DIR, "registro_maestro.xlsx")

# Cursos válidos, en orden
CURSOS = [
    "Pre-Kinder", "Kinder",
    "1°A", "1°B", "2°A", "2°B",
    "3°A", "3°B", "4°A", "4°B",
    "5°A", "5°B", "6°A", "6°B",
    "7°A", "7°B", "8°A", "8°B",
]

# Columnas del Excel maestro (segun usuario)
HEADERS = [
    "N°", "Curso", "Nombre Alumno", "Nombre Apoderado",
    "Correo Apoderado", "Teléfono 1", "Teléfono 2",
]

# Estilos para los encabezados
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Anchos de columna
COLUMN_WIDTHS = {
    "A": 6,   # N°
    "B": 14,  # Curso
    "C": 30,  # Nombre Alumno
    "D": 30,  # Nombre Apoderado
    "E": 35,  # Correo Apoderado
    "F": 16,  # Teléfono 1
    "G": 16,  # Teléfono 2
}


def ensure_data_dir():
    """se crea el directorio data/ si no existe."""
    os.makedirs(DATA_DIR, exist_ok=True)


def create_master_excel() -> str:
    """
    Crea el Excel maestro desde cero con todas las hojas de cursos.
    Retorna la ruta del archivo creado.
    """
    ensure_data_dir()
    wb = Workbook()

    # Eliminar la hoja por defecto
    default_sheet = wb.active
    wb.remove(default_sheet)

    for curso in CURSOS:
        ws = wb.create_sheet(title=curso)
        _setup_sheet_headers(ws)

    wb.save(MASTER_FILE)
    return MASTER_FILE


def _setup_sheet_headers(ws):
    """Configura los encabezados y formato de una hoja."""
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BORDER

    # Ajustar anchos de columna
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width


BACKUP_DIR = os.path.join(DATA_DIR, "backups")
MAX_BACKUPS = 10


def backup_master_excel():
    """
    Crea un backup del Excel maestro con timestamp antes de modificarlo.
    Ejemplo: registro_maestro_2026-03-17_03-50-32.xlsx
    Mantiene solo los últimos MAX_BACKUPS backups, eliminando los más antiguos.
    """
    if not os.path.exists(MASTER_FILE):
        return  # No hay nada que respaldar (primera vez)

    os.makedirs(BACKUP_DIR, exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    backup_name = f"registro_maestro_{timestamp}.xlsx"
    backup_path = os.path.join(BACKUP_DIR, backup_name)
    shutil.copy2(MASTER_FILE, backup_path)

    # Limpiar backups antiguos: mantener solo los últimos MAX_BACKUPS
    _cleanup_old_backups()

    return backup_path


def _cleanup_old_backups():
    """Elimina los backups más antiguos si hay más de MAX_BACKUPS."""
    backups = sorted(
        [
            f for f in os.listdir(BACKUP_DIR)
            if f.startswith("registro_maestro_") and f.endswith(".xlsx")
        ]
    )
    while len(backups) > MAX_BACKUPS:
        oldest = backups.pop(0)
        os.remove(os.path.join(BACKUP_DIR, oldest))



def load_or_create_master() -> str:
    """
    Carga el Excel maestro existente o lo crea si no existe.
    Retorna la ruta del archivo.
    """
    ensure_data_dir()
    if not os.path.exists(MASTER_FILE):
        return create_master_excel()
    return MASTER_FILE


def add_records(curso: str, records: list[dict]) -> str:
    """
    Agrega registros a la hoja del curso especificado en el Excel maestro.
    
    Argumentos:
        curso: Nombre del curso (debe estar en CURSOS)
        records: Lista de dicts con claves:
            nombre_alumno, nombre_apoderado, correo_apoderado, telefono_1, telefono_2
    
    Returns:
        Ruta del archivo Excel actualizado.
    """
    if curso not in CURSOS:
        raise ValueError(f"Curso no válido: {curso}. Cursos válidos: {CURSOS}")

    if not records:
        raise ValueError("No hay registros para agregar.")

    # Asegurar que el archivo existe
    load_or_create_master()

    # Hacer backup antes de modificar
    backup_master_excel()

    # Cargar el workbook existente (NO sobrescribir)
    wb = load_workbook(MASTER_FILE)

    # Obtener la hoja del curso
    if curso not in wb.sheetnames:
        # Si por alguna razón no existe la hoja, crearla
        ws = wb.create_sheet(title=curso)
        _setup_sheet_headers(ws)
    else:
        ws = wb[curso]

    # Determinar el próximo N° (autoincremental por hoja)
    last_row = ws.max_row
    if last_row <= 1:
        # Solo hay encabezados
        next_num = 1
    else:
        # Buscar el último N° usado
        last_num = 0
        for row in range(2, last_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value is not None:
                try:
                    last_num = max(last_num, int(cell_value))
                except (ValueError, TypeError):
                    pass
        next_num = last_num + 1

    # Agregar las filas nuevas
    for record in records:
        row_data = [
            next_num,
            curso,
            record.get("nombre_alumno") or "",
            record.get("nombre_apoderado") or "",
            record.get("correo_apoderado") or "",
            record.get("telefono_1") or "",
            record.get("telefono_2") or "",
        ]
        ws.append(row_data)

        # Aplicar bordes a la fila nueva
        current_row = ws.max_row
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = HEADER_BORDER
            cell.alignment = Alignment(horizontal="center" if col_idx <= 2 else "left")

        next_num += 1

    wb.save(MASTER_FILE)
    return MASTER_FILE


def get_master_path() -> str:
    """Retorna la ruta del Excel maestro, o None si no existe."""
    if os.path.exists(MASTER_FILE):
        return MASTER_FILE
    return None
